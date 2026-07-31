import json
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DetailView, DeleteView
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.shortcuts import redirect, get_object_or_404, render
from django.utils import timezone
from django.urls import reverse_lazy, reverse
from accounts.mixins import RoleRequiredMixin, ModalFormMixin, ModalDetailMixin, is_modal_request, modal_redirect
from django.http import JsonResponse
from scheduling.models import Session, SessionRequest
from scheduling.forms import SessionForm, SessionRequestForm
from scheduling.utils import generate_reservation_code, capacity_error, student_count_error
from labs.models import Lab
from notifications import services as notify_service

User = get_user_model()


def _official_schedule_conflict(lab, date, start_time, end_time, exclude_pk=None):
    """True if the given lab/date/time-range overlaps an already-APPROVED
    Session on the official schedule. Shared by request creation, request
    editing, and approval so 'is this slot free' always means the same
    thing everywhere it's asked."""
    qs = Session.objects.filter(
        lab=lab, date=date,
        start_time__lt=end_time, end_time__gt=start_time,
    )
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)
    return qs.exists()


def _slot_error(requester_type, lab, date, start_time, end_time, pcs_requested, student_count=0, exclude_pk=None):
    """The single gate used everywhere a request/session's lab+time slot
    needs checking: Instructor/Student/Group (including Class Roster
    reservations, which are just an Instructor request with a roster
    attached) keep the original all-or-nothing exclusive rule (any
    overlapping approved session blocks it) — but are now also validated
    against the lab's total PC count via scheduling.utils.student_count_error,
    since a class of more students than the lab has PCs for can't actually
    be seated. Walk-in/Override instead go through the PC-count capacity
    check (see scheduling.utils.capacity_error) so they can share a slot
    based on how many PCs are actually free. Returns an error message, or
    None if the slot/request is fine."""
    if requester_type in ('walk_in', 'override'):
        return capacity_error(requester_type, lab, date, start_time, end_time, pcs_requested, exclude_pk=exclude_pk)
    error = student_count_error(lab, student_count)
    if error:
        return error
    if _official_schedule_conflict(lab, date, start_time, end_time, exclude_pk=exclude_pk):
        return 'Time slot conflict — this laboratory is already booked for an overlapping time on that date. Choose a different laboratory or schedule.'
    return None


# ============================================================
# Schedule & Request hub — links into "View schedule" and
# "Manage requests"
# ============================================================
class ScheduleView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'scheduling/schedule.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['session_count'] = Session.objects.count()
        ctx['pending_count'] = SessionRequest.objects.filter(status='pending').count()
        return ctx


# ============================================================
# VIEW SCHEDULE FLOW
# Schedule and Request -> View Schedule -> Select Lab & Date
# -> Retrieve Schedule -> Schedule Found? -> Display / No result
# ============================================================
class ViewScheduleView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'scheduling/view_schedule.html'

    # Fixed color palette cycled per-lab so each lab keeps a consistent
    # color on the calendar regardless of how many labs exist.
    LAB_COLORS = ['#5b9dff', '#3ed6c4', '#a78bfa', '#f2a93b', '#ff5d5d', '#e879c9', '#7dd3fc', '#facc15']

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['pending_count'] = SessionRequest.objects.filter(status='pending').count()
        labs = list(Lab.objects.order_by('name'))
        ctx['labs'] = labs
        lab_id = self.request.GET.get('lab', '')
        date = self.request.GET.get('date', '')
        ctx['selected_lab'] = lab_id
        ctx['selected_date'] = date
        ctx['filtered'] = bool(lab_id or date)

        # Calendar shows the full range for the chosen lab (date is only
        # used to jump the calendar to that day, not to hide other days).
        sessions = Session.objects.select_related('instructor', 'lab').order_by('date', 'start_time')
        if lab_id:
            sessions = sessions.filter(lab_id=lab_id)
        ctx['schedule_found'] = sessions.exists()

        lab_color = {lab.pk: self.LAB_COLORS[i % len(self.LAB_COLORS)] for i, lab in enumerate(labs)}
        events = []
        for s in sessions:
            events.append({
                'id': s.pk,
                'title': f'{s.subject} — {s.lab.name}',
                'start': f'{s.date.isoformat()}T{s.start_time.isoformat(timespec="minutes")}',
                'end': f'{s.date.isoformat()}T{s.end_time.isoformat(timespec="minutes")}',
                'color': lab_color.get(s.lab_id, '#5b9dff'),
                'extendedProps': {
                    'lab': s.lab.name,
                    'subject': s.subject,
                    'time': f'{s.start_time.strftime("%I:%M %p").lstrip("0")}–{s.end_time.strftime("%I:%M %p").lstrip("0")}',
                    'requester': s.requester_name,
                    'requesterType': s.get_requester_type_display(),
                    'code': s.reservation_code or '',
                    'editUrl': reverse('session_edit', args=[s.pk]),
                    'deleteUrl': reverse('session_delete', args=[s.pk]),
                },
            })
        ctx['calendar_events_json'] = json.dumps(events)
        ctx['initial_date'] = date or timezone.localdate().isoformat()
        return ctx


def _retrieve_schedule(request):
    """Shared lookup used by both the view-schedule page and its exports."""
    lab_id = request.GET.get('lab')
    date = request.GET.get('date')
    qs = Session.objects.select_related('instructor', 'lab').order_by('start_time')
    if lab_id:
        qs = qs.filter(lab_id=lab_id)
    if date:
        qs = qs.filter(date=date)
    return qs, lab_id, date


# Action: Edit -> fix mistakes on an already-scheduled (approved) session
class SessionUpdateView(RoleRequiredMixin, ModalFormMixin, UpdateView):
    allowed_roles = ['admin', 'incharge']
    model = Session
    form_class = SessionForm
    template_name = 'scheduling/session_edit.html'
    context_object_name = 'session'

    def form_valid(self, form):
        error = _slot_error(
            form.instance.requester_type, form.instance.lab, form.instance.date,
            form.instance.start_time, form.instance.end_time, form.instance.pcs_requested,
            student_count=form.instance.student_count, exclude_pk=self.object.pk,
        )
        if error:
            form.add_error(None, error)
            return self.form_invalid(form)

        # Form validation (RequiresRegisteredAccountMixin) guarantees this
        # matches an active, admin-registered account before we ever get here.
        form.instance.instructor = form.cleaned_data['_matched_account']
        response = super().form_valid(form)
        messages.success(self.request, 'Schedule updated.')
        return response

    def get_success_url(self):
        return f"{reverse_lazy('view_schedule')}?lab={self.object.lab_id}&date={self.object.date}"


# Action: Delete -> remove an approved session from the official schedule
def delete_session(request, pk):
    session = get_object_or_404(Session, pk=pk)
    if request.method == 'POST':
        lab_id, date = session.lab_id, session.date
        session.delete()
        messages.success(request, 'Session removed from the schedule.')
        return redirect(f"{reverse_lazy('view_schedule')}?lab={lab_id}&date={date}")
    return redirect('view_schedule')


# Action: Export PDF (from the displayed schedule)
def export_schedule_pdf(request):
    sessions, lab_id, date = _retrieve_schedule(request)

    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    lab = Lab.objects.filter(pk=lab_id).first() if lab_id else None

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title='Lab Schedule')
    styles = getSampleStyleSheet()
    elements = [
        Paragraph('CompuLab — Lab Schedule', styles['Title']),
        Spacer(1, 6),
        Paragraph(f'Lab: {lab.name if lab else "All labs"} · Date: {date or "All dates"}', styles['Normal']),
        Paragraph(f'Generated {timezone.localtime(timezone.now()).strftime("%B %d, %Y %H:%M")}', styles['Normal']),
        Spacer(1, 16),
    ]

    data = [['Time', 'Class / subject', 'Requester', 'Lab']]
    for s in sessions:
        data.append([
            f'{s.start_time.strftime("%H:%M")}–{s.end_time.strftime("%H:%M")}',
            s.subject,
            s.requester_name or (s.instructor.get_full_name() if s.instructor else '—'),
            s.lab.name,
        ])
    if len(data) == 1:
        data.append(['—', 'No schedule found', '—', '—'])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#12171d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="lab_schedule.pdf"'
    return response


# Action: Export Excel (from the displayed schedule)
def export_schedule_excel(request):
    sessions, lab_id, date = _retrieve_schedule(request)

    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Lab Schedule'

    headers = ['Time', 'Class / subject', 'Requester', 'Lab', 'Date']
    ws.append(headers)
    header_fill = PatternFill(start_color='12171D', end_color='12171D', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill

    for s in sessions:
        ws.append([
            f'{s.start_time.strftime("%H:%M")}–{s.end_time.strftime("%H:%M")}',
            s.subject,
            s.requester_name or (s.instructor.get_full_name() if s.instructor else '—'),
            s.lab.name,
            s.date.strftime('%Y-%m-%d'),
        ])

    for col in ws.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="lab_schedule.xlsx"'
    return response


# ============================================================
# MANAGE REQUEST FLOW
# Schedule and Request -> Manage Request -> View Pending Requests
# -> Select Request -> Show Request Details -> Action
# ============================================================
class ManageRequestsView(RoleRequiredMixin, ListView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'scheduling/manage_requests.html'
    context_object_name = 'pending_requests'

    def get_queryset(self):
        return SessionRequest.objects.filter(
            status='pending'
        ).select_related('instructor', 'lab').order_by('date', 'start_time')


class RequestDetailView(RoleRequiredMixin, ModalDetailMixin, DetailView):
    allowed_roles = ['admin', 'incharge']
    model = SessionRequest
    template_name = 'scheduling/request_detail.html'
    context_object_name = 'req'


# Action: Edit -> fix mistakes on a still-pending request before Approve/Reject
class RequestUpdateView(RoleRequiredMixin, ModalFormMixin, UpdateView):
    allowed_roles = ['admin', 'incharge']
    model = SessionRequest
    form_class = SessionRequestForm
    template_name = 'scheduling/request_edit.html'
    context_object_name = 'req'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status != 'pending':
            messages.error(request, 'Only pending requests can be edited.')
            if is_modal_request(request):
                return JsonResponse({'success': True, 'redirect': reverse_lazy('request_detail', kwargs={'pk': self.object.pk}).__str__()})
            return redirect('request_detail', pk=self.object.pk)
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status != 'pending':
            messages.error(request, 'Only pending requests can be edited.')
            if is_modal_request(request):
                return JsonResponse({'success': True, 'redirect': reverse_lazy('request_detail', kwargs={'pk': self.object.pk}).__str__()})
            return redirect('request_detail', pk=self.object.pk)
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        error = _slot_error(
            form.instance.requester_type, form.instance.lab, form.instance.date,
            form.instance.start_time, form.instance.end_time, form.instance.pcs_requested,
            student_count=form.instance.student_count,
        )
        if error:
            form.add_error(None, error)
            return self.form_invalid(form)

        # Form validation (RequiresRegisteredAccountMixin) guarantees this
        # matches an active, admin-registered account before we ever get here.
        form.instance.instructor = form.cleaned_data['_matched_account']
        response = super().form_valid(form)
        messages.success(self.request, 'Request updated.')
        return response

    def get_success_url(self):
        return reverse_lazy('request_detail', kwargs={'pk': self.object.pk})


# Action: Approve -> Add to Official Schedule -> Return to SS view (pending list)
def approve_request(request, pk):
    req = get_object_or_404(SessionRequest, pk=pk)
    if request.method == 'POST':
        error = _slot_error(
            req.requester_type, req.lab, req.date, req.start_time, req.end_time, req.pcs_requested,
            student_count=req.student_count,
        )
        if error:
            if 'conflict' in error.lower():
                notify_service.notify_reservation_conflict(
                    req.lab,
                    f'Approving "{req.subject}" ({req.date} {req.start_time.strftime("%H:%M")}–'
                    f'{req.end_time.strftime("%H:%M")}) in {req.lab.name} conflicts with an existing booking.',
                )
            messages.error(request, error)
            if is_modal_request(request):
                return JsonResponse({'success': True, 'redirect': reverse('request_detail', kwargs={'pk': pk})})
            return redirect('request_detail', pk=pk)

        code = generate_reservation_code()
        req.reservation_code = code
        req.status = 'approved'
        req.save()

        # create the actual session — adds it to the official schedule,
        # carrying the reservation code the requester will use at the lab PC
        Session.objects.create(
            lab=req.lab,
            instructor=req.instructor,
            requester_type=req.requester_type,
            requester_name=req.requester_name,
            requester_id_number=req.requester_id_number,
            roster=req.roster,
            reservation_code=code,
            subject=req.subject,
            date=req.date,
            start_time=req.start_time,
            end_time=req.end_time,
            student_count=req.student_count,
            pcs_requested=req.pcs_requested,
        )

        # notify the linked account, if there is one, plus the relevant
        # Lab In-Charge/Admins — walk-in/override approvals get their own
        # event since there's usually no linked instructor account for them.
        notify_service.notify_reservation_approved(req, code)
        if req.requester_type in ('walk_in', 'override'):
            notify_service.notify_walkin_override_approved(req, code)
        messages.success(
            request,
            f'Request approved and added to the official schedule. Reservation code: {code} — give this to {req.requester_name}.'
        )
        if is_modal_request(request):
            return JsonResponse({'success': True, 'redirect': reverse('manage_requests')})
        return redirect('manage_requests')
    return redirect('manage_requests')


# Action: Reject -> Request Rejected -> Return to Schedule & request
def decline_request(request, pk):
    req = get_object_or_404(SessionRequest, pk=pk)
    if request.method == 'POST':
        req.status = 'declined'
        req.save()
        notify_service.notify_reservation_declined(req)
        messages.warning(request, 'Request rejected.')
        if is_modal_request(request):
            return JsonResponse({'success': True, 'redirect': reverse('schedule')})
        return redirect('schedule')
    return redirect('schedule')


# Action: Export PDF for a single request -> Return to Schedule & request
def export_request_pdf(request, pk):
    req = get_object_or_404(SessionRequest, pk=pk)

    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title='Lab Request')
    styles = getSampleStyleSheet()
    elements = [
        Paragraph('CompuLab — Lab Request', styles['Title']),
        Spacer(1, 6),
        Paragraph(f'Generated {timezone.localtime(timezone.now()).strftime("%B %d, %Y %H:%M")}', styles['Normal']),
        Spacer(1, 16),
    ]

    data = [
        ['Requester', f'{req.requester_name} ({req.get_requester_type_display()})'],
        ['ID number', req.requester_id_number],
        ['Lab', req.lab.name],
        ['Subject', req.subject],
        ['Date', req.date.strftime('%B %d, %Y')],
        ['Time', f'{req.start_time.strftime("%H:%M")}–{req.end_time.strftime("%H:%M")}'],
        ['Students', str(req.student_count)],
        ['Status', req.get_status_display()],
        ['Reservation code', req.reservation_code or '— (not yet approved)'],
        ['Notes', req.notes or '—'],
    ]

    table = Table(data, colWidths=[110, 360])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f0e8')),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="request_{req.pk}.pdf"'
    return response


# ============================================================
# Other schedule pages (unchanged)
# ============================================================

# lab schedule — incharge sees their assigned lab only
class LabScheduleView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'scheduling/lab_schedule.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        assigned_lab = self.request.user.assigned_lab
        ctx['sessions'] = Session.objects.filter(
            lab=assigned_lab
        ).order_by('date', 'start_time')
        ctx['lab'] = assigned_lab
        return ctx


# instructor sees only their own schedule
class InstructorScheduleView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['instructor']
    template_name = 'scheduling/instructor_schedule.html'

    def get_context_data(self, **kwargs):
        from scheduling.models import ClassRoster
        ctx = super().get_context_data(**kwargs)
        ctx['my_sessions'] = Session.objects.filter(
            instructor=self.request.user
        ).order_by('date', 'start_time')
        ctx['my_requests'] = SessionRequest.objects.filter(
            instructor=self.request.user
        ).order_by('-created_at')
        ctx['my_rosters'] = ClassRoster.objects.filter(
            instructor=self.request.user
        ).order_by('-created_at')
        return ctx


# Admin/incharge logs a reservation on behalf of an instructor or student
# who requested a lab in person or verbally — no system account needed.
# Still goes through the normal Approve/Reject step once encoded.
class RequestCreateView(RoleRequiredMixin, ModalFormMixin, CreateView):
    allowed_roles = ['admin', 'incharge']
    model = SessionRequest
    form_class = SessionRequestForm
    template_name = 'scheduling/request_new.html'

    def form_valid(self, form):
        # Per the reservation workflow: check the lab's availability for
        # this date/time BEFORE saving — a logged request that already
        # collides with an approved session on the official schedule isn't
        # useful to anyone (it'll only fail later at approval, after the
        # Admin/In-Charge has already told the requester it went through).
        # This is a same-lab check against Session (see approve_request for
        # why it's re-checked again there too: something else may have been
        # approved into this slot in the time between logging and approving
        # this particular request).
        error = _slot_error(
            form.instance.requester_type, form.instance.lab, form.instance.date,
            form.instance.start_time, form.instance.end_time, form.instance.pcs_requested,
            student_count=form.instance.student_count,
        )
        if error:
            if 'conflict' in error.lower():
                notify_service.notify_reservation_conflict(
                    form.instance.lab,
                    f'A new request for "{form.instance.subject}" ({form.instance.date} '
                    f'{form.instance.start_time.strftime("%H:%M")}–{form.instance.end_time.strftime("%H:%M")}) '
                    f'in {form.instance.lab.name} conflicts with an existing booking.',
                )
            form.add_error(None, error)
            return self.form_invalid(form)

        form.instance.status = 'pending'
        # Only accounts registered by an Admin/Lab In-Charge (under Users)
        # may hold a reservation — SessionRequestForm's
        # RequiresRegisteredAccountMixin already rejected this submission
        # if no matching active account exists, so this is guaranteed here.
        form.instance.instructor = form.cleaned_data['_matched_account']
        response = super().form_valid(form)
        notify_service.notify_reservation_submitted(self.object)
        messages.success(self.request, 'Reservation logged. Review the details and approve or reject it.')
        return response

    def get_success_url(self):
        return reverse_lazy('request_detail', kwargs={'pk': self.object.pk})


# ============================================================
# CLASS ROSTER — list of enrolled students (ID + name) per class,
# used by the Attendance Report to show real present/absent names
# instead of PC-login-only estimates.
# ============================================================
class RosterListView(RoleRequiredMixin, ListView):
    allowed_roles = ['admin', 'incharge', 'instructor']
    template_name = 'scheduling/roster_list.html'
    context_object_name = 'rosters'
    paginate_by = 10

    def _base_queryset(self):
        from scheduling.models import ClassRoster
        qs = ClassRoster.objects.select_related('lab', 'instructor')
        if self.request.user.role == 'instructor':
            qs = qs.filter(instructor=self.request.user)
        return qs

    def get_queryset(self):
        qs = self._base_queryset().order_by('status', 'name')
        q = self.request.GET.get('q', '').strip()
        semester = self.request.GET.get('semester', '').strip()
        course = self.request.GET.get('course', '').strip()
        status = self.request.GET.get('status', '').strip()
        if q:
            qs = qs.filter(
                Q(name__icontains=q) | Q(course_code__icontains=q) |
                Q(subject__icontains=q) | Q(section__icontains=q) |
                Q(instructor__first_name__icontains=q) | Q(instructor__last_name__icontains=q)
            )
        if semester:
            qs = qs.filter(semester=semester)
        if course:
            qs = qs.filter(course_code=course)
        if status:
            qs = qs.filter(status=status)
        approval = self.request.GET.get('approval', '').strip()
        if approval:
            qs = qs.filter(approval_status=approval)
        return qs

    def get_context_data(self, **kwargs):
        from django.db.models import Count
        from scheduling.models import ClassRoster
        ctx = super().get_context_data(**kwargs)
        base_qs = self._base_queryset()
        ctx['total_rosters'] = base_qs.count()
        ctx['total_students'] = base_qs.aggregate(n=Count('students'))['n'] or 0
        ctx['active_rosters'] = base_qs.filter(status='active').count()
        ctx['archived_rosters'] = base_qs.filter(status='inactive').count()
        ctx['pending_rosters'] = base_qs.filter(approval_status='pending').count()
        ctx['course_options'] = (
            base_qs.exclude(course_code='').order_by('course_code')
            .values_list('course_code', flat=True).distinct()
        )
        ctx['semester_choices'] = ClassRoster.SEMESTER_CHOICES
        ctx['approval_choices'] = ClassRoster.APPROVAL_CHOICES
        ctx['selected_q'] = self.request.GET.get('q', '')
        ctx['selected_semester'] = self.request.GET.get('semester', '')
        ctx['selected_course'] = self.request.GET.get('course', '')
        ctx['selected_status'] = self.request.GET.get('status', '')
        ctx['selected_approval'] = self.request.GET.get('approval', '')
        return ctx


class RosterCreateView(RoleRequiredMixin, ModalFormMixin, CreateView):
    allowed_roles = ['admin', 'incharge', 'instructor']
    template_name = 'scheduling/roster_form.html'

    def get_form_class(self):
        from scheduling.forms import ClassRosterForm
        return ClassRosterForm

    def get_initial(self):
        initial = super().get_initial()
        if self.request.user.role == 'instructor':
            initial['instructor'] = self.request.user.pk
        return initial

    def form_valid(self, form):
        response = super().form_valid(form)

        # Admin/Lab In-Charge already hold approval authority, so a roster
        # they create themselves doesn't need to sit in a review queue
        # waiting on... themselves. An Instructor-created roster stays at
        # the model default ('pending') and must be reviewed.
        if self.request.user.role in ('admin', 'incharge'):
            self.object.approval_status = 'approved'
            self.object.save(update_fields=['approval_status'])

        if self.object.approval_status != 'approved':
            from notifications import services as notify_service
            notify_service.notify_roster_submitted(self.object)
            messages.success(
                self.request,
                f'Roster "{self.object.name}" created and is Pending approval by an Admin or Lab '
                f'In-Charge. Its schedule will be added to the official calendar once approved.'
            )
        elif self.object.has_full_schedule():
            created, skipped = self.object.generate_sessions()
            if created:
                messages.success(
                    self.request,
                    f'Roster "{self.object.name}" created and approved. {len(created)} scheduled session(s) '
                    f'were automatically added to the official schedule. Each reservation code will be sent '
                    f'to the instructor\'s notification panel 1 hour before that session starts.'
                )
            else:
                messages.warning(
                    self.request,
                    f'Roster "{self.object.name}" created, but no sessions could be auto-scheduled for '
                    f'the chosen date range — see below.'
                )
            if skipped:
                messages.warning(
                    self.request,
                    f'{len(skipped)} date(s) in the range could not be auto-scheduled (lab already booked) '
                    f'and were skipped — add those manually if still needed.'
                )
        else:
            messages.success(self.request, f'Roster "{self.object.name}" created. Add students on the next page.')
        return response

    def get_success_url(self):
        return reverse_lazy('roster_detail', kwargs={'pk': self.object.pk})


class RosterDetailView(RoleRequiredMixin, ModalDetailMixin, DetailView):
    allowed_roles = ['admin', 'incharge', 'instructor']
    template_name = 'scheduling/roster_detail.html'
    context_object_name = 'roster'

    def get_queryset(self):
        from scheduling.models import ClassRoster
        return ClassRoster.objects.select_related('lab', 'instructor')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['students'] = self.object.students.select_related('student').all()
        ctx['can_review_roster'] = self.request.user.role in ('admin', 'incharge')

        now = timezone.localtime()
        today, now_time = now.date(), now.time()
        sessions = list(self.object.sessions.select_related('lab').order_by('-date', '-start_time')[:20])
        for s in sessions:
            if s.date < today or (s.date == today and s.end_time <= now_time):
                s.reservation_status = 'completed'
            elif s.date == today and s.start_time <= now_time < s.end_time:
                s.reservation_status = 'ongoing'
            else:
                s.reservation_status = 'scheduled'
        ctx['linked_sessions'] = sessions
        return ctx


def roster_generate_sessions(request, pk):
    """Manual re-run of ClassRoster.generate_sessions() for an existing
    roster — lets Admin/In-Charge/instructor backfill or top up scheduled
    sessions (e.g. after extending the validity period) without having to
    re-save the whole roster form. Safe to click repeatedly: dates already
    generated are skipped, not duplicated."""
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster
    if request.user.role not in ('admin', 'incharge', 'instructor'):
        raise PermissionDenied
    roster = get_object_or_404(ClassRoster, pk=pk)
    if request.method == 'POST':
        pruned = roster.prune_stale_future_sessions()
        if pruned:
            messages.info(
                request,
                f'{pruned} upcoming session(s) that no longer match this roster\'s schedule were '
                f'removed (past sessions and any with a recorded check-in were left untouched).'
            )
        if roster.approval_status != 'approved':
            messages.error(
                request,
                'This roster must be Approved by an Admin or Lab In-Charge before its schedule can '
                'be added to the official calendar.'
            )
        elif not roster.has_full_schedule():
            messages.error(
                request,
                'Set a laboratory, meeting day(s)/time, and validity period (from/until) on this '
                'roster before generating sessions.'
            )
        else:
            created, skipped = roster.generate_sessions()
            if created:
                messages.success(
                    request,
                    f'{len(created)} scheduled session(s) added to the official schedule. Each '
                    f'reservation code will be sent to the instructor\'s notification panel 1 hour '
                    f'before that session starts.'
                )
            else:
                messages.info(request, 'No new sessions to generate — the schedule is already fully rolled out.')
            if skipped:
                messages.warning(
                    request,
                    f'{len(skipped)} date(s) could not be auto-scheduled (lab already booked) and were skipped.'
                )
    return modal_redirect(request, 'roster_detail', pk=roster.pk)


def roster_archive(request, pk):
    """Toggle a roster between active and archived (inactive). Kept as a
    quick one-click action separate from the full edit form — the actual
    student/session records are untouched either way."""
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster
    if request.user.role not in ('admin', 'incharge', 'instructor'):
        raise PermissionDenied
    roster = get_object_or_404(ClassRoster, pk=pk)
    if request.method == 'POST':
        roster.status = 'inactive' if roster.status == 'active' else 'active'
        roster.save(update_fields=['status'])
        messages.success(
            request,
            f'Roster "{roster.name}" archived.' if roster.status == 'inactive'
            else f'Roster "{roster.name}" reactivated.'
        )
    return modal_redirect(request, 'roster_detail', pk=roster.pk)


# Action: Approve -> generate the roster's official schedule -> back to roster detail
def roster_approve(request, pk):
    """Only an Admin or Lab In-Charge may approve a pending roster. Its
    schedule is re-checked for conflicts at approval time (defensive —
    another roster/session could have been approved in the meantime) and,
    if clear, the roster is rolled onto the official calendar the same way
    RosterCreateView/RosterUpdateView already do for an Approved roster."""
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster
    from scheduling.utils import roster_schedule_conflicts
    from notifications import services as notify_service

    if request.user.role not in ('admin', 'incharge'):
        raise PermissionDenied
    roster = get_object_or_404(ClassRoster, pk=pk)
    if request.method == 'POST':
        if roster.approval_status != 'pending':
            messages.error(request, 'Only a Pending roster can be approved.')
            return modal_redirect(request, 'roster_detail', pk=pk)

        conflicts = roster_schedule_conflicts(
            lab=roster.lab, instructor=roster.instructor, section=roster.section,
            schedule_days=roster.schedule_days, start_time=roster.schedule_start_time,
            end_time=roster.schedule_end_time, valid_from=roster.schedule_valid_from,
            valid_until=roster.schedule_valid_until, exclude_pk=roster.pk,
        )
        if conflicts:
            first = conflicts[0]
            messages.error(
                request,
                f'Cannot approve — {first["kind"]} conflict with "{first["roster"].name}". '
                f'Edit this roster to a different available time before approving.'
            )
            return modal_redirect(request, 'roster_detail', pk=pk)

        roster.approval_status = 'approved'
        roster.save(update_fields=['approval_status'])
        notify_service.notify_roster_approved(roster)

        if roster.has_full_schedule():
            created, skipped = roster.generate_sessions()
            if created:
                messages.success(
                    request,
                    f'Roster "{roster.name}" approved. {len(created)} scheduled session(s) were added '
                    f'to the official schedule.'
                )
            else:
                messages.success(request, f'Roster "{roster.name}" approved.')
            if skipped:
                messages.warning(
                    request,
                    f'{len(skipped)} date(s) in the range could not be auto-scheduled (lab already booked) '
                    f'and were skipped — add those manually if still needed.'
                )
        else:
            messages.success(request, f'Roster "{roster.name}" approved.')
    return modal_redirect(request, 'roster_detail', pk=pk)


# Action: Reject -> instructor notified, roster stays off the official schedule
def roster_reject(request, pk):
    """Only an Admin or Lab In-Charge may reject a pending roster. The
    instructor is notified and can edit + resubmit — see
    RosterUpdateView.form_valid, which resets a Rejected roster back to
    Pending as soon as it's edited."""
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster
    from notifications import services as notify_service

    if request.user.role not in ('admin', 'incharge'):
        raise PermissionDenied
    roster = get_object_or_404(ClassRoster, pk=pk)
    if request.method == 'POST':
        if roster.approval_status != 'pending':
            messages.error(request, 'Only a Pending roster can be rejected.')
            return modal_redirect(request, 'roster_detail', pk=pk)

        reason = (request.POST.get('reason') or '').strip()
        roster.approval_status = 'rejected'
        roster.rejection_reason = reason
        roster.save(update_fields=['approval_status', 'rejection_reason'])
        notify_service.notify_roster_rejected(roster)
        messages.warning(request, f'Roster "{roster.name}" rejected.')
    return modal_redirect(request, 'roster_detail', pk=pk)


class RosterUpdateView(RoleRequiredMixin, ModalFormMixin, UpdateView):
    allowed_roles = ['admin', 'incharge', 'instructor']
    template_name = 'scheduling/roster_form.html'

    def get_form_class(self):
        from scheduling.forms import ClassRosterForm
        return ClassRosterForm

    def get_queryset(self):
        from scheduling.models import ClassRoster
        return ClassRoster.objects.all()

    def form_valid(self, form):
        was_rejected = self.object.approval_status == 'rejected'
        response = super().form_valid(form)

        if was_rejected:
            # Editing a rejected roster is a resubmission — send it back
            # into the review queue rather than leaving it Rejected.
            self.object.approval_status = 'pending'
            self.object.rejection_reason = ''
            self.object.save(update_fields=['approval_status', 'rejection_reason'])
            from notifications import services as notify_service
            notify_service.notify_roster_submitted(self.object)
            messages.success(
                self.request,
                f'Roster "{self.object.name}" updated and resubmitted — it is Pending approval again.'
            )
        else:
            messages.success(self.request, f'Roster "{self.object.name}" updated.')

        pruned = self.object.prune_stale_future_sessions()
        if pruned:
            messages.info(
                self.request,
                f'{pruned} upcoming session(s) from the roster\'s previous schedule no longer match '
                f'and were removed from the official schedule (past sessions and any with a recorded '
                f'check-in were left untouched).'
            )
        if self.object.approval_status == 'approved' and self.object.has_full_schedule():
            created, skipped = self.object.generate_sessions()
            if created:
                messages.success(
                    self.request,
                    f'{len(created)} newly-scheduled session(s) were added to the official schedule. '
                    f'Each reservation code will be sent to the instructor\'s notification panel 1 hour '
                    f'before that session starts.'
                )
            if skipped:
                messages.warning(
                    self.request,
                    f'{len(skipped)} date(s) in the range could not be auto-scheduled (lab already booked) '
                    f'and were skipped — add those manually if still needed.'
                )
        return response

    def get_success_url(self):
        return reverse_lazy('roster_detail', kwargs={'pk': self.object.pk})


class RosterDeleteView(RoleRequiredMixin, ModalFormMixin, DeleteView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'scheduling/roster_confirm_delete.html'
    context_object_name = 'roster'
    success_url = reverse_lazy('roster_list')

    def get_queryset(self):
        from scheduling.models import ClassRoster
        return ClassRoster.objects.all()

    def form_valid(self, form):
        messages.success(self.request, f'Roster "{self.object.name}" deleted.')
        return super().form_valid(form)


def roster_check_availability(request):
    """Live-lookup endpoint for the roster form: given a lab/instructor/
    section, returns every OTHER active pending-or-approved roster's
    weekly schedule for each of them, so the form can show 'occupied'
    slots next to the day/time picker before the instructor even submits —
    per the rule that the system should surface already-occupied time
    slots up front, not just reject a conflicting submission after the fact."""
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster
    from django.http import JsonResponse

    if request.user.role not in ('admin', 'incharge', 'instructor'):
        raise PermissionDenied

    lab_id = request.GET.get('lab')
    instructor_id = request.GET.get('instructor')
    section = (request.GET.get('section') or '').strip()
    exclude_pk = request.GET.get('exclude')

    qs = ClassRoster.objects.filter(
        status='active', approval_status__in=('pending', 'approved'),
    ).exclude(schedule_days='')
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)

    def _serialize(roster_qs):
        out = []
        for r in roster_qs.select_related('instructor'):
            out.append({
                'name': r.name,
                'days': [ClassRoster.DAY_LABELS.get(d, d) for d in r.schedule_days.split(',') if d],
                'start': r.schedule_start_time.strftime('%I:%M %p').lstrip('0') if r.schedule_start_time else '',
                'end': r.schedule_end_time.strftime('%I:%M %p').lstrip('0') if r.schedule_end_time else '',
                'valid_from': r.schedule_valid_from.strftime('%b %d, %Y') if r.schedule_valid_from else '',
                'valid_until': r.schedule_valid_until.strftime('%b %d, %Y') if r.schedule_valid_until else '',
                'approval_status': r.approval_status,
            })
        return out

    data = {'lab_schedule': [], 'instructor_schedule': [], 'section_schedule': []}
    if lab_id:
        data['lab_schedule'] = _serialize(qs.filter(lab_id=lab_id))
    if instructor_id:
        data['instructor_schedule'] = _serialize(qs.filter(instructor_id=instructor_id))
    if section:
        data['section_schedule'] = _serialize(qs.filter(section=section))
    return JsonResponse(data)


def roster_search_students(request, pk):
    """Live-search endpoint for the 'Add student' box on the roster detail
    page. Only searches role='student' accounts (registered by an Admin
    under Users) and excludes students already on this roster — matching
    the 'only officially registered students, no duplicates' requirement."""
    from django.http import JsonResponse
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster

    if not request.user.is_authenticated or request.user.role not in ('admin', 'incharge', 'instructor'):
        raise PermissionDenied
    roster = get_object_or_404(ClassRoster, pk=pk)

    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse({'results': []})

    already_on_roster = set(roster.students.exclude(student__isnull=True).values_list('student_id', flat=True))
    students = User.objects.filter(
        Q(role='student'), Q(is_active=True),
    ).filter(
        Q(first_name__icontains=q) | Q(last_name__icontains=q) |
        Q(id_number__icontains=q)
    ).exclude(pk__in=already_on_roster).order_by('first_name', 'last_name')[:15]

    return JsonResponse({
        'results': [
            {
                'id': s.pk,
                'id_number': s.id_number or s.display_name,
                'full_name': s.display_name,
                'department': s.department_display,
                'year_level': s.year_level_display,
            }
            for s in students
        ]
    })


def roster_add_student(request, pk):
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster, RosterStudent
    from scheduling.forms import RosterAddStudentForm
    if not request.user.is_authenticated or request.user.role not in ('admin', 'incharge', 'instructor'):
        raise PermissionDenied
    roster = get_object_or_404(ClassRoster, pk=pk)
    if request.method == 'POST':
        form = RosterAddStudentForm(request.POST)
        if form.is_valid():
            student = form.cleaned_data['student']
            _, created = RosterStudent.objects.get_or_create(roster=roster, student=student)
            if created:
                messages.success(request, f'Added {student.display_name} to the roster.')
            else:
                messages.error(request, f'{student.display_name} is already on this roster.')
        else:
            messages.error(request, 'Could not add that student — please search and select again.')
    return modal_redirect(request, 'roster_detail', pk=pk)


def roster_remove_student(request, pk, student_pk):
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster, RosterStudent
    if not request.user.is_authenticated or request.user.role not in ('admin', 'incharge', 'instructor'):
        raise PermissionDenied
    roster = get_object_or_404(ClassRoster, pk=pk)
    if request.method == 'POST':
        RosterStudent.objects.filter(pk=student_pk, roster=roster).delete()
        messages.success(request, 'Student removed from roster.')
    return modal_redirect(request, 'roster_detail', pk=pk)


def roster_import_students(request, pk):
    """'Import Students (Excel/CSV)' for a Class Roster — bulk version of
    the search-and-add box above. Scoped to one Department per upload:
    only already-registered, active student accounts in that Department
    get matched and added; nothing is auto-created here (unlike the Users
    import), since a roster may only contain officially registered accounts."""
    from django.core.exceptions import PermissionDenied
    from scheduling.models import ClassRoster, RosterStudent
    from scheduling.forms import RosterImportStudentsForm
    from accounts.imports import read_rows
    from accounts.constants import DEPARTMENT_CHOICES, department_name

    if not request.user.is_authenticated or request.user.role not in ('admin', 'incharge', 'instructor'):
        raise PermissionDenied
    roster = get_object_or_404(ClassRoster, pk=pk)

    results = None
    if request.method == 'POST':
        form = RosterImportStudentsForm(request.POST, request.FILES)
        if form.is_valid():
            department = form.cleaned_data['department']
            try:
                rows = read_rows(form.cleaned_data['file'])
            except ValueError as e:
                form.add_error('file', str(e))
                rows = None

            if rows is not None:
                if not rows:
                    form.add_error('file', 'No data rows found in that file.')
                else:
                    added, skipped = [], []
                    already_on_roster = set(
                        roster.students.exclude(student__isnull=True).values_list('student_id', flat=True)
                    )
                    for i, row in enumerate(rows, start=2):
                        id_number = row.get('id_number', '')
                        if not id_number:
                            skipped.append({'row': i, 'id_number': '—', 'reason': 'Missing ID Number.'})
                            continue
                        student = User.objects.filter(
                            role='student', id_number=id_number, department=department,
                        ).first()
                        if not student:
                            skipped.append({
                                'row': i, 'id_number': id_number,
                                'reason': f'No registered student with this ID in {department_name(department)}.',
                            })
                            continue
                        if not student.is_active:
                            skipped.append({'row': i, 'id_number': id_number, 'reason': 'That account is deactivated.'})
                            continue
                        if student.pk in already_on_roster:
                            skipped.append({'row': i, 'id_number': id_number, 'reason': 'Already on this roster.'})
                            continue
                        RosterStudent.objects.create(roster=roster, student=student)
                        already_on_roster.add(student.pk)
                        added.append({'row': i, 'id_number': id_number, 'name': student.display_name})

                    if added:
                        messages.success(request, f'Added {len(added)} student{"s" if len(added) != 1 else ""} to the roster.')
                    if skipped:
                        messages.warning(request, f'{len(skipped)} row{"s were" if len(skipped) != 1 else " was"} skipped — see details below.')
                    results = {'added': added, 'skipped': skipped, 'department': department_name(department)}
                    form = RosterImportStudentsForm()
    else:
        form = RosterImportStudentsForm()

    return render(request, 'scheduling/roster_import.html', {
        'roster': roster, 'form': form, 'results': results, 'department_choices': DEPARTMENT_CHOICES,
    })
