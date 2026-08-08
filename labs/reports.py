"""
Reporting & Analytics — Lab Utilization, Instructor Usage, Attendance, and
Maintenance reports. Each report is: a filter form (GET params) -> computed
data -> an on-screen table -> optional PDF/Excel export using the same
filtered queryset/computation.
"""
from datetime import datetime, timedelta, date as date_cls

from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView

from accounts.mixins import RoleRequiredMixin
from accounts.models import ActivityLog
from labs.models import Lab, PC, InventoryItem, MaintenanceLog, EquipmentIssue

User = get_user_model()


def _scoped_lab_id(request, lab_id):
    """Lab In-Charge accounts are always scoped to their own assigned lab,
    no matter what the 'lab' filter says — an incharge can't page through
    other labs' reports just by editing the querystring."""
    if request.user.role == 'incharge' and request.user.assigned_lab_id:
        return str(request.user.assigned_lab_id)
    return lab_id


def _visible_labs(request):
    """Labs an incharge is allowed to pick from in a filter dropdown —
    just their own; admins see everything."""
    if request.user.role == 'incharge' and request.user.assigned_lab_id:
        return Lab.objects.filter(pk=request.user.assigned_lab_id)
    return Lab.objects.order_by('name')


def _require_report_access(request):
    if not request.user.is_authenticated:
        raise PermissionDenied
    if request.user.role not in ('admin', 'incharge'):
        raise PermissionDenied


def _log_report(request, report_type, fmt):
    """Record that this report was generated, so 'Recent Generated Reports'
    on the analytics page has real history to show."""
    from labs.models import ReportLog
    lab = None
    if request.user.role == 'incharge':
        lab = request.user.assigned_lab
    elif request.GET.get('lab'):
        lab = Lab.objects.filter(pk=request.GET.get('lab')).first()
    try:
        ReportLog.objects.create(
            report_type=report_type,
            format=fmt,
            generated_by=request.user,
            lab=lab,
            params=request.GET.urlencode()[:255],
        )
    except Exception:
        # Logging history should never break the actual export.
        pass


# ---------------------------------------------------------------------------
# shared helpers
# ---------------------------------------------------------------------------

def _parse_date(value, fallback):
    if not value:
        return fallback
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return fallback


def _default_range():
    today = timezone.now().date()
    return today - timedelta(days=30), today


def _session_hours(session):
    """Duration of a Session in hours (float), guarding against bad data."""
    start = datetime.combine(session.date, session.start_time)
    end = datetime.combine(session.date, session.end_time)
    seconds = (end - start).total_seconds()
    return max(seconds, 0) / 3600


def _peak_hours(sessions):
    """Returns a list of (hour, session_count) sorted by count desc, for the
    hour-of-day buckets that at least one session overlaps."""
    buckets = {}
    for s in sessions:
        start_hour = s.start_time.hour
        end_hour = s.end_time.hour if s.end_time.minute or s.end_time.second else s.end_time.hour - 1
        end_hour = max(end_hour, start_hour)
        for h in range(start_hour, end_hour + 1):
            buckets[h] = buckets.get(h, 0) + 1
    return sorted(buckets.items(), key=lambda pair: (-pair[1], pair[0]))


def _fmt_hour(h):
    suffix = 'AM' if h < 12 else 'PM'
    display = h % 12
    if display == 0:
        display = 12
    return f'{display}:00 {suffix}'


# ---------------------------------------------------------------------------
# Report 1 — Lab Utilization
# ---------------------------------------------------------------------------

def _lab_utilization_data(request):
    from scheduling.models import Session

    default_from, default_to = _default_range()
    date_from = _parse_date(request.GET.get('date_from'), default_from)
    date_to = _parse_date(request.GET.get('date_to'), default_to)
    lab_id = _scoped_lab_id(request, request.GET.get('lab', ''))

    sessions = Session.objects.filter(date__gte=date_from, date__lte=date_to)
    if lab_id:
        sessions = sessions.filter(lab_id=lab_id)
    sessions = sessions.select_related('lab').order_by('lab__name', 'date', 'start_time')

    num_days = (date_to - date_from).days + 1
    labs = Lab.objects.filter(pk=lab_id) if lab_id else Lab.objects.all()

    rows = []
    for lab in labs.order_by('name'):
        lab_sessions = [s for s in sessions if s.lab_id == lab.id]
        total_hours = sum(_session_hours(s) for s in lab_sessions)
        open_hours = datetime.combine(date_cls.min, lab.closing_time) - datetime.combine(date_cls.min, lab.opening_time)
        capacity_hours = (open_hours.total_seconds() / 3600) * num_days
        utilization_pct = (total_hours / capacity_hours * 100) if capacity_hours > 0 else 0
        peak = _peak_hours(lab_sessions)
        rows.append({
            'lab': lab,
            'total_sessions': len(lab_sessions),
            'total_hours': round(total_hours, 1),
            'capacity_hours': round(capacity_hours, 1),
            'utilization_pct': round(utilization_pct, 1),
            'peak_hours': [(_fmt_hour(h), c) for h, c in peak[:3]],
        })

    return {
        'date_from': date_from, 'date_to': date_to, 'selected_lab': lab_id,
        'all_labs': _visible_labs(request), 'rows': rows,
    }


class LabUtilizationReportView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'labs/reports/lab_utilization.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_lab_utilization_data(self.request))
        return ctx


def export_lab_utilization(request):
    _require_report_access(request)
    data = _lab_utilization_data(request)
    fmt = request.GET.get('format', 'pdf')
    title = 'Lab Utilization Report'
    period = f"{data['date_from']} to {data['date_to']}"
    headers = ['Lab', 'Sessions', 'Total Hours', 'Capacity Hours', 'Utilization %', 'Peak Hours']
    body = [[
        r['lab'].name, r['total_sessions'], r['total_hours'], r['capacity_hours'],
        f"{r['utilization_pct']}%",
        ', '.join(f'{h} ({c})' for h, c in r['peak_hours']) or '—',
    ] for r in data['rows']]
    if fmt == 'excel':
        _log_report(request, 'lab_utilization', 'excel')
        return _export_excel(title, period, headers, body, 'lab_utilization_report.xlsx')
    _log_report(request, 'lab_utilization', 'pdf')
    return _export_pdf(title, period, headers, body, 'lab_utilization_report.pdf')


# ---------------------------------------------------------------------------
# Report 1b — Equipment Status
# ---------------------------------------------------------------------------

def _equipment_status_data(request):
    default_from, default_to = _default_range()
    date_from = _parse_date(request.GET.get('date_from'), default_from)
    date_to = _parse_date(request.GET.get('date_to'), default_to)
    lab_id = _scoped_lab_id(request, request.GET.get('lab', ''))

    items = InventoryItem.objects.select_related('lab').filter(
        last_checked__gte=date_from, last_checked__lte=date_to
    )
    if lab_id:
        items = items.filter(lab_id=lab_id)
    items = items.order_by('lab__name', 'name')

    rows = []
    for item in items:
        rows.append({
            'item': item,
            'open_issues': item.issues.filter(status='open').count(),
        })

    summary = {
        'total': len(rows),
        'operational': sum(1 for r in rows if r['item'].status == 'operational'),
        'under_repair': sum(1 for r in rows if r['item'].status == 'under_repair'),
        'retired': sum(1 for r in rows if r['item'].status == 'retired'),
    }

    return {
        'date_from': date_from, 'date_to': date_to, 'selected_lab': lab_id,
        'all_labs': _visible_labs(request), 'rows': rows, 'summary': summary,
    }


class EquipmentStatusReportView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'labs/reports/equipment_status.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_equipment_status_data(self.request))
        return ctx


def export_equipment_status(request):
    _require_report_access(request)
    data = _equipment_status_data(request)
    fmt = request.GET.get('format', 'pdf')
    title = 'Equipment Status Report'
    period = f"{data['date_from']} to {data['date_to']}"
    headers = ['Equipment', 'Category', 'Lab', 'Condition', 'Status', 'Quantity', 'Last Checked', 'Open Issues']
    body = [[
        r['item'].name, r['item'].category, r['item'].lab.name,
        r['item'].get_condition_display(), r['item'].get_status_display(),
        r['item'].quantity, r['item'].last_checked or '—', r['open_issues'],
    ] for r in data['rows']]
    if fmt == 'excel':
        _log_report(request, 'equipment_status', 'excel')
        return _export_excel(title, period, headers, body, 'equipment_status_report.xlsx')
    _log_report(request, 'equipment_status', 'pdf')
    return _export_pdf(title, period, headers, body, 'equipment_status_report.pdf')


# ---------------------------------------------------------------------------
# Report 2 — Instructor Usage
# ---------------------------------------------------------------------------

def _instructor_usage_data(request):
    from scheduling.models import Session

    default_from, default_to = _default_range()
    date_from = _parse_date(request.GET.get('date_from'), default_from)
    date_to = _parse_date(request.GET.get('date_to'), default_to)
    instructor_id = request.GET.get('instructor', '')

    sessions = Session.objects.filter(
        date__gte=date_from, date__lte=date_to, instructor__isnull=False
    ).select_related('lab', 'instructor')
    if request.user.role == 'incharge' and request.user.assigned_lab_id:
        sessions = sessions.filter(lab_id=request.user.assigned_lab_id)
    if instructor_id:
        sessions = sessions.filter(instructor_id=instructor_id)

    by_instructor = {}
    for s in sessions:
        entry = by_instructor.setdefault(s.instructor, {'labs': set(), 'hours': 0.0, 'sessions': 0})
        entry['labs'].add(s.lab.name)
        entry['hours'] += _session_hours(s)
        entry['sessions'] += 1

    rows = [{
        'instructor': instr,
        'labs_used': sorted(v['labs']),
        'total_hours': round(v['hours'], 1),
        'num_sessions': v['sessions'],
    } for instr, v in by_instructor.items()]
    rows.sort(key=lambda r: r['instructor'].display_name)

    return {
        'date_from': date_from, 'date_to': date_to, 'selected_instructor': instructor_id,
        'instructors': User.objects.filter(role='instructor').order_by('first_name', 'last_name'),
        'rows': rows,
    }


class InstructorUsageReportView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'labs/reports/instructor_usage.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_instructor_usage_data(self.request))
        return ctx


def export_instructor_usage(request):
    _require_report_access(request)
    data = _instructor_usage_data(request)
    fmt = request.GET.get('format', 'pdf')
    title = 'Instructor Usage Report'
    period = f"{data['date_from']} to {data['date_to']}"
    headers = ['Instructor', 'Laboratories Used', 'Total Hours', 'Number of Sessions']
    body = [[
        r['instructor'].display_name,
        ', '.join(r['labs_used']) or '—',
        r['total_hours'], r['num_sessions'],
    ] for r in data['rows']]
    if fmt == 'excel':
        _log_report(request, 'instructor_usage', 'excel')
        return _export_excel(title, period, headers, body, 'instructor_usage_report.xlsx')
    _log_report(request, 'instructor_usage', 'pdf')
    return _export_pdf(title, period, headers, body, 'instructor_usage_report.pdf')


# ---------------------------------------------------------------------------
# Report 3 — Attendance
# ---------------------------------------------------------------------------

def _norm_id(id_number):
    """Canonical form of an ID number for matching only (never display).
    Check-in matches an ID case-insensitively, but the roster/present/
    absent comparison below used to compare the raw, as-typed ID against
    the roster's stored ID with exact-case string equality — so a student
    who checked in with different casing/whitespace than how their ID was
    entered on the roster (e.g. a CSV import) was recorded as present
    under their typed ID AND still listed as absent under their roster
    ID. Normalizing both sides before comparing fixes that false
    "Absent"."""
    return (id_number or '').strip().casefold()


def _attendance_data(request):
    from scheduling.models import Session

    lab_id = _scoped_lab_id(request, request.GET.get('lab', ''))
    day = _parse_date(request.GET.get('day'), timezone.now().date())

    sessions = Session.objects.filter(date=day).select_related('lab', 'roster')
    if lab_id:
        sessions = sessions.filter(lab_id=lab_id)
    sessions = sessions.order_by('lab__name', 'start_time')

    rows = []
    for s in sessions:
        start_dt = datetime.combine(day, s.start_time)
        end_dt = datetime.combine(day, s.end_time)
        logs = ActivityLog.objects.filter(
            action='pc_unlock', pc__lab_id=s.lab_id,
            created_at__gte=timezone.make_aware(start_dt) if timezone.is_naive(start_dt) else start_dt,
            created_at__lte=timezone.make_aware(end_dt) if timezone.is_naive(end_dt) else end_dt,
        ).order_by('created_at')

        roster_by_id = {}
        if s.roster_id:
            # Keyed by normalized ID so a roster ID and a differently-cased/
            # spaced typed ID from the same student are treated as one
            # person instead of two (present under one, absent under the
            # other).
            roster_by_id = {_norm_id(rs.id_number): (rs.id_number, rs.full_name) for rs in s.roster.students.all()}

        # Walk-in/Override check-ins (people using a PC during this slot who
        # aren't the requester, a roster member, or a counted group member —
        # see SessionCheckIn/checkin_type) get their own section below, not
        # mixed into this session's roster present/absent accounting.
        walk_in_checkins = {
            c.id_number: c for c in
            s.check_ins.filter(checkin_type__in=('walk_in', 'override')).select_related('student', 'pc')
        }
        walk_in_ids_norm = {_norm_id(i) for i in walk_in_checkins}

        present = []
        seen_ids = set()
        for log in logs:
            norm = _norm_id(log.target_identifier)
            if norm in seen_ids or norm in walk_in_ids_norm:
                continue
            seen_ids.add(norm)
            # Prefer the roster's own name; if this person isn't on the
            # roster (e.g. a Group booking's members, who aren't named
            # anywhere), fall back to the resolved account from
            # SessionCheckIn rather than the requester's name — this used
            # to show the reservation's requester for every group member,
            # which was wrong.
            checkin = s.check_ins.filter(id_number=log.target_identifier).select_related('student').first()
            roster_entry = roster_by_id.get(norm)
            name = (
                (roster_entry[1] if roster_entry else None)
                or (checkin.student.display_name if checkin and checkin.student else None)
                or s.requester_name or '—'
            )
            present.append({
                'id_number': log.target_identifier,
                'name': name,
                'pc': log.pc.pc_id if log.pc else '—',
                'time': log.created_at,
            })

        # A session that hasn't started yet can't have "absent" students —
        # nobody has had the chance to check in. Without this check, every
        # roster member on an upcoming session was marked absent the moment
        # the report was generated, well before class time.
        now = timezone.localtime()
        session_started = s.date < now.date() or (s.date == now.date() and s.start_time <= now.time())

        if not session_started:
            absent = [] if s.roster_id else None
            not_started = True
            expected = len(roster_by_id) if s.roster_id else (
                s.pcs_requested if s.requester_type in ('walk_in', 'override') else (s.student_count or 0)
            )
        else:
            not_started = False
            if s.roster_id:
                # roster attached — absent list has real names, not just a count
                absent = [
                    {'id_number': id_number, 'full_name': full_name}
                    for norm, (id_number, full_name) in roster_by_id.items()
                    if norm not in seen_ids
                ]
                expected = len(roster_by_id)
            else:
                absent = None  # no roster — names aren't knowable, only a count
                expected = s.pcs_requested if s.requester_type in ('walk_in', 'override') else (s.student_count or 0)

        # Walk-in / Override transactions: people who used a PC during this
        # session's slot without being the requester, a roster member, or a
        # counted group member. Shown as their own section (with their real
        # resolved name/account) instead of being folded into "present"
        # under the requester's name.
        walk_ins = [
            {
                'id_number': c.id_number,
                'name': (c.student.display_name) if c.student else c.id_number,
                'role': c.student.get_role_display() if c.student else '—',
                'checkin_type': c.get_checkin_type_display(),
                'pc': c.pc.pc_id if c.pc else '—',
                'time': c.checked_in_at,
            }
            for c in walk_in_checkins.values()
        ]

        present_count = len(present)
        if not_started:
            absent_count = 0
        else:
            absent_count = len(absent) if absent is not None else max(expected - present_count, 0)
        rows.append({
            'session': s,
            'has_roster': bool(s.roster_id),
            'not_started': not_started,
            'present': present,
            'absent': absent,
            'walk_ins': walk_ins,
            'present_count': present_count,
            'absent_count': absent_count,
            'expected_count': expected,
        })

    return {
        'day': day, 'selected_lab': lab_id,
        'all_labs': _visible_labs(request), 'rows': rows,
    }


class AttendanceReportView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'labs/reports/attendance.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_attendance_data(self.request))
        return ctx


def export_attendance(request):
    _require_report_access(request)
    data = _attendance_data(request)
    fmt = request.GET.get('format', 'pdf')
    title = 'Attendance Report'
    period = f"{data['day']}"
    headers = ['Lab', 'Subject', 'Time', 'Present', 'Absent', 'Time Logs', 'Absent Names (if roster attached)', 'Walk-in / Override']
    body = []
    for r in data['rows']:
        s = r['session']
        time_logs = '; '.join(f"{p['name']} ({p['id_number']}) @ {timezone.localtime(p['time']).strftime('%H:%M')}" for p in r['present']) or '—'
        if r.get('not_started'):
            absent_names = 'Session has not started yet'
        elif r['has_roster']:
            absent_names = '; '.join(f"{a['full_name']} ({a['id_number']})" for a in r['absent']) or 'None'
        else:
            absent_names = 'No roster attached — estimate only'
        walk_in_text = '; '.join(
            f"{w['name']} ({w['id_number']}, {w['role']}) — {w['checkin_type']} @ {timezone.localtime(w['time']).strftime('%H:%M')}"
            for w in r['walk_ins']
        ) or '—'
        body.append([
            s.lab.name, s.subject, f'{s.start_time.strftime("%H:%M")}-{s.end_time.strftime("%H:%M")}',
            r['present_count'], r['absent_count'], time_logs, absent_names, walk_in_text,
        ])
    if fmt == 'excel':
        _log_report(request, 'attendance', 'excel')
        return _export_excel(title, period, headers, body, 'attendance_report.xlsx')
    _log_report(request, 'attendance', 'pdf')
    return _export_pdf(title, period, headers, body, 'attendance_report.pdf')


# ---------------------------------------------------------------------------
# Report 4 — Maintenance
# ---------------------------------------------------------------------------

def _maintenance_data(request):
    default_from, default_to = _default_range()
    date_from = _parse_date(request.GET.get('date_from'), default_from)
    date_to = _parse_date(request.GET.get('date_to'), default_to)
    equipment_id = request.GET.get('equipment', '')

    logs = MaintenanceLog.objects.filter(
        maintenance_date__gte=date_from, maintenance_date__lte=date_to
    ).select_related('equipment', 'assigned_technician')
    issues = EquipmentIssue.objects.filter(
        date_reported__date__gte=date_from, date_reported__date__lte=date_to
    ).select_related('equipment')
    damaged = InventoryItem.objects.filter(condition__in=['faulty', 'needs_check']) | \
        InventoryItem.objects.filter(status='retired')
    damaged = damaged.distinct().select_related('lab')

    all_equipment = InventoryItem.objects.order_by('name')
    if request.user.role == 'incharge' and request.user.assigned_lab_id:
        logs = logs.filter(equipment__lab_id=request.user.assigned_lab_id)
        issues = issues.filter(equipment__lab_id=request.user.assigned_lab_id)
        damaged = damaged.filter(lab_id=request.user.assigned_lab_id)
        all_equipment = all_equipment.filter(lab_id=request.user.assigned_lab_id)

    if equipment_id:
        logs = logs.filter(equipment_id=equipment_id)
        issues = issues.filter(equipment_id=equipment_id)
        damaged = damaged.filter(pk=equipment_id)

    scheduled = logs.filter(completed=False).order_by('maintenance_date')
    performed = logs.filter(completed=True).order_by('-maintenance_date')

    return {
        'date_from': date_from, 'date_to': date_to, 'selected_equipment': equipment_id,
        'all_equipment': all_equipment,
        'scheduled': scheduled, 'performed': performed, 'issues': issues, 'damaged': damaged,
    }


class MaintenanceReportView(RoleRequiredMixin, TemplateView):
    allowed_roles = ['admin', 'incharge']
    template_name = 'labs/reports/maintenance.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_maintenance_data(self.request))
        return ctx


def export_maintenance(request):
    _require_report_access(request)
    data = _maintenance_data(request)
    fmt = request.GET.get('format', 'pdf')
    title = 'Maintenance Report'
    period = f"{data['date_from']} to {data['date_to']}"

    headers = ['Section', 'Equipment', 'Date', 'Technician / Notes', 'Status']
    body = []
    for log in data['scheduled']:
        body.append(['Scheduled', log.equipment.name, str(log.maintenance_date),
                      log.assigned_technician.get_full_name() if log.assigned_technician else '—', 'Pending'])
    for log in data['performed']:
        body.append(['Repair performed', log.equipment.name, str(log.maintenance_date),
                      (log.completion_notes or log.notes)[:80], 'Completed'])
    for item in data['damaged']:
        body.append(['Lost/Damaged', item.name, str(item.last_checked or '—'),
                      f'{item.get_condition_display()} — {item.lab.name}', item.get_status_display()])

    if fmt == 'excel':
        _log_report(request, 'maintenance', 'excel')
        return _export_excel(title, period, headers, body, 'maintenance_report.xlsx')
    _log_report(request, 'maintenance', 'pdf')
    return _export_pdf(title, period, headers, body, 'maintenance_report.pdf')


# ---------------------------------------------------------------------------
# Reporting & Analytics overview — quick summary export (the header's
# Export PDF / Export Excel buttons). Uses the same lab-utilization
# computation as the main report, just condensed to one summary table.
# ---------------------------------------------------------------------------

def export_analytics_overview(request):
    _require_report_access(request)
    data = _lab_utilization_data(request)
    fmt = request.GET.get('format', 'pdf')
    title = 'Reporting & Analytics — Overview'
    period = f"{data['date_from']} to {data['date_to']}"
    headers = ['Lab', 'Sessions', 'Total Hours', 'Utilization %']
    body = [[r['lab'].name, r['total_sessions'], r['total_hours'], f"{r['utilization_pct']}%"] for r in data['rows']]
    if fmt == 'excel':
        return _export_excel(title, period, headers, body, 'analytics_overview.xlsx')
    return _export_pdf(title, period, headers, body, 'analytics_overview.pdf')


# ---------------------------------------------------------------------------
# generic PDF / Excel export helpers
# ---------------------------------------------------------------------------

def export_pc_status_report(request):
    """Exports the currently filtered PC Power Status list (same lab/status/
    search filters as the on-screen page) as a PDF or Excel snapshot."""
    _require_report_access(request)

    lab_id = request.GET.get('lab', '')
    status_filter = request.GET.get('status', '')
    q = request.GET.get('q', '').strip()

    pcs = PC.objects.select_related('lab', 'current_user')
    if request.user.role == 'incharge' and request.user.assigned_lab_id:
        pcs = pcs.filter(lab_id=request.user.assigned_lab_id)
    if lab_id:
        pcs = pcs.filter(lab_id=lab_id)
    if status_filter:
        pcs = pcs.filter(status=status_filter)
    if q:
        from django.db.models import Q
        pcs = pcs.filter(
            Q(pc_id__icontains=q) |
            Q(current_user__first_name__icontains=q) |
            Q(current_user__last_name__icontains=q) |
            Q(current_user__id_number__icontains=q) |
            Q(current_user__email__icontains=q)
        )
    pcs = pcs.order_by('lab__name', 'pc_id')

    title = 'PC Power Status Report'
    period = f'As of {timezone.localtime(timezone.now()).strftime("%B %d, %Y %H:%M")}'
    headers = ['PC ID', 'Lab', 'IP Address', 'Status', 'Current User', 'Last Active']
    body = []
    for pc in pcs:
        user_name = pc.current_user.display_name if pc.current_user else '—'
        last_active = timezone.localtime(pc.last_active).strftime('%b %d, %Y %H:%M') if pc.last_active else '—'
        body.append([pc.pc_id, pc.lab.name, pc.ip_address or '—', pc.get_status_display(), user_name, last_active])

    fmt = request.GET.get('format', 'pdf')
    if fmt == 'excel':
        return _export_excel(title, period, headers, body, 'pc_power_status_report.xlsx')
    return _export_pdf(title, period, headers, body, 'pc_power_status_report.pdf')


# ---------------------------------------------------------------------------
# generic PDF / Excel export helpers
# ---------------------------------------------------------------------------

def _export_pdf(title, period, headers, rows, filename):
    from io import BytesIO
    from django.conf import settings
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable,
    )

    INK = colors.HexColor('#12171d')
    TEAL = colors.HexColor('#3ed6c4')
    TEXT_DIM = colors.HexColor('#4d5761')
    BORDER = colors.HexColor('#d8dde1')
    ROW_ALT = colors.HexColor('#f2f7f6')

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter), title=title,
        topMargin=0.45 * inch, bottomMargin=0.45 * inch,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    school_style = ParagraphStyle('school', parent=styles['Normal'], fontSize=9,
                                   textColor=TEXT_DIM, spaceAfter=1, leading=11)
    title_style = ParagraphStyle('titleStyle', parent=styles['Title'], textColor=INK,
                                  fontSize=17, leading=20, spaceAfter=2, spaceBefore=0)
    meta_style = ParagraphStyle('meta', parent=styles['Normal'], fontSize=8.5,
                                 textColor=TEXT_DIM, leading=11)

    header_text = [
        Paragraph('DANAO TECHNOLOGICAL COLLEGE', school_style),
        Paragraph(f'CompuLab — {title}', title_style),
        Paragraph(
            f'Period: {period} &nbsp;&nbsp;&middot;&nbsp;&nbsp; '
            f'Generated {timezone.localtime(timezone.now()).strftime("%B %d, %Y %H:%M")}',
            meta_style,
        ),
    ]

    logo_path = settings.BASE_DIR / 'static' / 'img' / 'logo.jpg'
    if logo_path.exists():
        logo_cell = Image(str(logo_path), width=0.72 * inch, height=0.72 * inch)
    else:
        logo_cell = ''

    header_table = Table([[logo_cell, header_text]], colWidths=[0.95 * inch, None])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('LEFTPADDING', (1, 0), (1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    elements = [
        header_table,
        Spacer(1, 8),
        HRFlowable(width='100%', thickness=1.6, color=TEAL, spaceAfter=14),
    ]

    data = [headers] + [[str(cell) for cell in row] for row in rows] if rows else [headers, ['No data for this filter.'] + [''] * (len(headers) - 1)]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), INK),
        ('TEXTCOLOR', (0, 0), (-1, 0), TEAL),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setStrokeColor(BORDER)
        canvas.line(0.5 * inch, 0.35 * inch, doc_.pagesize[0] - 0.5 * inch, 0.35 * inch)
        canvas.setFont('Helvetica', 7.5)
        canvas.setFillColor(TEXT_DIM)
        canvas.drawString(0.5 * inch, 0.22 * inch, 'CompuLab — Danao Technological College')
        canvas.drawRightString(doc_.pagesize[0] - 0.5 * inch, 0.22 * inch, f'Page {doc_.page}')
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_excel(title, period, headers, rows, filename):
    from io import BytesIO
    from django.conf import settings
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    INK = '12171D'
    TEAL = '3ED6C4'
    TEXT_DIM = '8B98A3'
    ROW_ALT = 'F2F7F6'
    BORDER_COLOR = 'D8DDE1'
    thin = Side(style='thin', color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    n_cols = max(len(headers), 1)
    last_col_letter = get_column_letter(n_cols)
    text_col = 1

    logo_path = settings.BASE_DIR / 'static' / 'img' / 'logo.jpg'
    if logo_path.exists():
        img = XLImage(str(logo_path))
        img.width = 58
        img.height = 58
        ws.add_image(img, 'A1')
        text_col = 2  # leave column A clear for the logo

    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 15
    ws.column_dimensions['A'].width = 9

    ws.cell(row=1, column=text_col, value='DANAO TECHNOLOGICAL COLLEGE').font = Font(size=9, color=TEXT_DIM)
    ws.cell(row=2, column=text_col, value=f'CompuLab — {title}').font = Font(size=15, bold=True, color=INK)
    ws.cell(row=3, column=text_col,
            value=(f'Period: {period}   ·   Generated '
                   f'{timezone.localtime(timezone.now()).strftime("%B %d, %Y %H:%M")}')
            ).font = Font(size=9, color=TEXT_DIM)

    header_row = 5
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=INK)
        c.border = border
        c.alignment = Alignment(vertical='center')

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    if rows:
        for r_i, row in enumerate(rows):
            excel_row = header_row + 1 + r_i
            row_fill = PatternFill('solid', fgColor=ROW_ALT) if r_i % 2 == 1 else None
            for c_i, val in enumerate(row, start=1):
                cell = ws.cell(row=excel_row, column=c_i, value=val)
                cell.border = border
                if row_fill:
                    cell.fill = row_fill
        last_row = header_row + len(rows)
    else:
        ws.cell(row=header_row + 1, column=1, value='No data for this filter.')
        last_row = header_row + 1

    ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{last_row}'

    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        length = max(
            (len(str(ws.cell(row=r, column=col_idx).value))
             for r in range(header_row, last_row + 1)
             if ws.cell(row=r, column=col_idx).value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response