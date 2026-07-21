"""
Parsing for the 'Import Students (Excel/CSV)' flows.

Both User & Role Management (create student accounts under a department)
and Class Roster (bulk-add existing students to a roster) upload the same
kind of file, so the row-reading logic lives here once.
"""
import csv
import io

from accounts.constants import year_level_choices_for

# Recognized header names, lower-cased, mapped to our internal field name.
# Accepts a few common spellings/spacings so an admin's existing sheet is
# more likely to work without edits first.
_HEADER_ALIASES = {
    'id number': 'id_number', 'id_number': 'id_number', 'idnumber': 'id_number', 'id': 'id_number',
    'first name': 'first_name', 'first_name': 'first_name', 'firstname': 'first_name',
    'last name': 'last_name', 'last_name': 'last_name', 'lastname': 'last_name',
    'year level': 'year_level', 'year_level': 'year_level', 'yearlevel': 'year_level', 'year': 'year_level',
    'section': 'section',
    'username': 'username',
}


def _normalize_row(headers, raw_row):
    row = {}
    for header, value in zip(headers, raw_row):
        key = _HEADER_ALIASES.get((header or '').strip().lower())
        if key:
            row[key] = ('' if value is None else str(value)).strip()
    return row


def read_rows(uploaded_file):
    """Yields one dict per data row: {'id_number', 'first_name', 'last_name',
    'year_level', 'section', 'username'} (missing columns simply absent).
    Raises ValueError with a human-readable message if the file can't be
    read at all."""
    name = (uploaded_file.name or '').lower()

    if name.endswith('.csv'):
        try:
            text = uploaded_file.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            text = uploaded_file.read().decode('latin-1')
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return []
        headers = rows[0]
        return [_normalize_row(headers, r) for r in rows[1:] if any((c or '').strip() for c in r)]

    if name.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError('Excel support (openpyxl) isn\'t installed on the server.')
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows_iter))
        except StopIteration:
            return []
        out = []
        for raw_row in rows_iter:
            if any(c not in (None, '') for c in raw_row):
                out.append(_normalize_row(headers, raw_row))
        return out

    raise ValueError('Unsupported file type — please upload a .csv or .xlsx file.')


def parse_year_level(raw):
    """'1', '1st Year', 'Year 1', 1.0 -> 1. Returns None if it can't be
    confidently parsed as a whole number."""
    if raw in (None, ''):
        return None
    digits = ''.join(ch for ch in str(raw) if ch.isdigit())
    return int(digits) if digits else None


def validate_row(row, department):
    """Returns a list of error strings for one parsed row (empty = valid).
    `department` is the department code the whole import is scoped to."""
    errors = []
    if not row.get('id_number'):
        errors.append('Missing ID Number.')
    if not row.get('first_name') and not row.get('last_name'):
        errors.append('Missing First/Last Name.')

    year_level_raw = row.get('year_level')
    if year_level_raw:
        year_level = parse_year_level(year_level_raw)
        valid_values = {v for v, _ in year_level_choices_for(department)}
        if year_level not in valid_values:
            errors.append(f'Year Level "{year_level_raw}" isn\'t offered by the selected department.')
    return errors
